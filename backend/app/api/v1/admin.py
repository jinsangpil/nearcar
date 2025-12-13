"""
운영자 API 엔드포인트
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, Dict
from datetime import date
from io import BytesIO

from app.core.database import get_db
from app.core.dependencies import require_role, require_admin_only, require_admin_or_staff
from app.schemas.admin import (
    PricePolicyCreateRequest,
    InspectionAssignRequest,
    SettlementCalculateRequest
)
from app.schemas.settlement import (
    SettlementListResponse,
    SettlementDetailResponse,
    SettlementSummaryResponse,
    SettlementStatusUpdateRequest,
    SettlementBulkUpdateRequest
)
from app.services.settlement_service import SettlementService
from app.schemas.price_policy import (
    PricePolicyResponse,
    PricePolicyListResponse,
    PricePolicyUpdateRequest
)
from app.services.price_policy_service import PricePolicyService
from app.schemas.service_region import (
    ServiceRegionCreateRequest,
    ServiceRegionUpdateRequest,
    ServiceRegionResponse,
    ServiceRegionListResponse
)
from app.services.service_region_service import ServiceRegionService
from app.schemas.vehicle_master import (
    VehicleMasterCreateRequest,
    VehicleMasterUpdateRequest,
    VehicleMasterResponse,
    VehicleMasterListResponse,
    VehicleMasterSyncRequest,
    VehicleMasterSyncResponse
)
from app.schemas.manufacturer import (
    ManufacturerCreateRequest,
    ManufacturerUpdateRequest,
    ManufacturerResponse,
    ManufacturerListResponse
)
from app.schemas.vehicle_model import (
    VehicleModelCreateRequest,
    VehicleModelUpdateRequest,
    VehicleModelResponse,
    VehicleModelListResponse,
    VehicleModelSyncRequest,
    VehicleModelSyncResponse
)
from app.schemas.user import (
    UserCreateRequest,
    UserUpdateRequest,
    UserResponse,
    UserListResponse,
    UserLevelUpdateRequest,
    UserCommissionUpdateRequest,
    UserRoleUpdateRequest,
    UserStatusUpdateRequest
)
from app.schemas.vehicle import StandardResponse
from app.schemas.package import (
    PackageCreateRequest,
    PackageUpdateRequest,
    PackageResponse,
    PackageListResponse
)
from app.services.admin_service import AdminService
from app.services.user_service import UserService
from app.services.package_service import PackageService
from app.services.vehicle_master_service import VehicleMasterService
from app.services.manufacturer_service import ManufacturerService
from app.services.vehicle_model_service import VehicleModelService
from app.models.user import User
from app.schemas.review import (
    ReviewResponse,
    ReviewListResponse,
    ReviewUpdateRequest
)
from app.services.review_service import ReviewService
from app.schemas.faq import (
    FAQCreateRequest,
    FAQUpdateRequest,
    FAQResponse,
    FAQListResponse
)
from app.services.faq_service import FAQService
import uuid

router = APIRouter(prefix="/admin", tags=["운영자"])


@router.get("/dashboard/stats", response_model=StandardResponse)
async def get_dashboard_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    대시보드 통계 API
    
    주요 지표 및 추이 데이터를 반환합니다.
    관리자 권한 필요.
    """
    try:
        result = await AdminService.get_dashboard_stats(db=db)
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"대시보드 통계 조회 중 오류가 발생했습니다: {str(e)}"
        )


# ==================== 차량 마스터 관리 API ====================
@router.post("/vehicles/master", response_model=StandardResponse)
async def create_vehicle_master(
    request: VehicleMasterCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 마스터 생성 API
    
    관리자 권한 필요.
    """
    try:
        new_master = await VehicleMasterService.create_vehicle_master(
            db=db,
            origin=request.origin,
            manufacturer=request.manufacturer,
            model_group=request.model_group,
            model_detail=request.model_detail,
            vehicle_class=request.vehicle_class,
            start_year=request.start_year,
            end_year=request.end_year,
            is_active=request.is_active
        )
        return StandardResponse(success=True, data=new_master)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 생성 중 오류 발생: {str(e)}")


@router.get("/vehicles/master/{master_id}", response_model=StandardResponse)
async def get_vehicle_master_detail(
    master_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    차량 마스터 상세 조회 API
    
    관리자/직원 권한 필요.
    """
    try:
        master_uuid = uuid.UUID(master_id)
        master = await VehicleMasterService.get_vehicle_master(db, master_uuid)
        if not master:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 마스터를 찾을 수 없습니다.")
        return StandardResponse(success=True, data=master)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="유효하지 않은 차량 마스터 ID 형식입니다.")
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 조회 중 오류 발생: {str(e)}")


@router.patch("/vehicles/master/{master_id}", response_model=StandardResponse)
async def update_vehicle_master(
    master_id: str,
    request: VehicleMasterUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 마스터 수정 API
    
    관리자 권한 필요.
    """
    try:
        master_uuid = uuid.UUID(master_id)
        updated_master = await VehicleMasterService.update_vehicle_master(
            db=db,
            master_id=master_uuid,
            origin=request.origin,
            manufacturer=request.manufacturer,
            model_group=request.model_group,
            model_detail=request.model_detail,
            vehicle_class=request.vehicle_class,
            start_year=request.start_year,
            end_year=request.end_year,
            is_active=request.is_active
        )
        if not updated_master:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 마스터를 찾을 수 없습니다.")
        return StandardResponse(success=True, data=updated_master)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 수정 중 오류 발생: {str(e)}")


@router.delete("/vehicles/master/{master_id}", response_model=StandardResponse)
async def delete_vehicle_master(
    master_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 마스터 삭제 API (soft delete)
    
    관리자 권한 필요.
    """
    try:
        master_uuid = uuid.UUID(master_id)
        success = await VehicleMasterService.delete_vehicle_master(db, master_uuid)
        if not success:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 마스터를 찾을 수 없습니다.")
        return StandardResponse(success=True, data={"message": "차량 마스터가 성공적으로 삭제되었습니다."})
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 삭제 중 오류 발생: {str(e)}")


@router.get("/vehicles/master", response_model=StandardResponse)
async def list_vehicle_masters(
    origin: Optional[str] = Query(None, description="국산/수입 필터 (domestic, imported)"),
    manufacturer: Optional[str] = Query(None, description="제조사 필터"),
    vehicle_class: Optional[str] = Query(None, description="차량 등급 필터"),
    search: Optional[str] = Query(None, description="검색어 (제조사, 모델명)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    차량 마스터 목록 조회 API
    
    관리자/직원 권한 필요.
    """
    try:
        masters_data = await VehicleMasterService.list_vehicle_masters(
            db=db,
            origin=origin,
            manufacturer=manufacturer,
            vehicle_class=vehicle_class,
            search=search,
            page=page,
            limit=limit
        )
        return StandardResponse(success=True, data=masters_data)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 목록 조회 중 오류 발생: {str(e)}")


@router.post("/vehicles/master/sync", response_model=StandardResponse)
async def sync_vehicle_masters(
    request: VehicleMasterSyncRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 마스터 일괄 동기화 API
    
    스크래핑 데이터를 일괄 동기화합니다.
    관리자 권한 필요.
    """
    try:
        sync_data = [item.model_dump() for item in request.data]
        result = await VehicleMasterService.sync_vehicle_masters(db, sync_data)
        return StandardResponse(success=True, data=result)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 마스터 동기화 중 오류 발생: {str(e)}")


# ==================== 제조사 관리 API ====================
@router.post("/manufacturers", response_model=StandardResponse)
async def create_manufacturer(
    request: ManufacturerCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    새 제조사를 생성합니다.
    관리자 권한 필요.
    """
    try:
        new_manufacturer = await ManufacturerService.create_manufacturer(
            db=db,
            name=request.name,
            origin=request.origin,
            is_active=request.is_active
        )
        return StandardResponse(success=True, data=new_manufacturer)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"제조사 생성 중 오류 발생: {str(e)}")


@router.get("/manufacturers/{manufacturer_id}", response_model=StandardResponse)
async def get_manufacturer_detail(
    manufacturer_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    특정 제조사 상세 정보를 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        manufacturer_uuid = uuid.UUID(manufacturer_id)
        manufacturer = await ManufacturerService.get_manufacturer(db, manufacturer_uuid)
        if not manufacturer:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="제조사를 찾을 수 없습니다.")
        return StandardResponse(success=True, data=manufacturer)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="유효하지 않은 제조사 ID 형식입니다.")
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"제조사 조회 중 오류 발생: {str(e)}")


@router.patch("/manufacturers/{manufacturer_id}", response_model=StandardResponse)
async def update_manufacturer(
    manufacturer_id: str,
    request: ManufacturerUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    제조사 정보를 업데이트합니다.
    관리자 권한 필요.
    """
    try:
        manufacturer_uuid = uuid.UUID(manufacturer_id)
        updated_manufacturer = await ManufacturerService.update_manufacturer(
            db=db,
            manufacturer_id=manufacturer_uuid,
            name=request.name,
            origin=request.origin,
            is_active=request.is_active
        )
        if not updated_manufacturer:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="제조사를 찾을 수 없습니다.")
        return StandardResponse(success=True, data=updated_manufacturer)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"제조사 업데이트 중 오류 발생: {str(e)}")


@router.delete("/manufacturers/{manufacturer_id}", response_model=StandardResponse)
async def delete_manufacturer(
    manufacturer_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    제조사를 삭제합니다 (soft delete).
    관리자 권한 필요.
    """
    try:
        manufacturer_uuid = uuid.UUID(manufacturer_id)
        success = await ManufacturerService.delete_manufacturer(db, manufacturer_uuid)
        if not success:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="제조사를 찾을 수 없습니다.")
        return StandardResponse(success=True, data={"message": "제조사가 성공적으로 삭제되었습니다."})
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"제조사 삭제 중 오류 발생: {str(e)}")


@router.get("/manufacturers", response_model=StandardResponse)
async def list_manufacturers(
    origin: Optional[str] = Query(None, description="국산/수입 필터 (domestic, imported)"),
    search: Optional[str] = Query(None, description="제조사명 검색"),
    is_active: Optional[bool] = Query(None, description="활성화 여부 필터"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    제조사 목록을 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        manufacturers_data = await ManufacturerService.list_manufacturers(
            db=db,
            origin=origin,
            search=search,
            is_active=is_active,
            page=page,
            limit=limit
        )
        return StandardResponse(success=True, data=manufacturers_data)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"제조사 목록 조회 중 오류 발생: {str(e)}")


# ==================== 차량 모델 관리 API ====================
@router.post("/vehicle-models", response_model=StandardResponse)
async def create_vehicle_model(
    request: VehicleModelCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    새 차량 모델을 생성합니다.
    관리자 권한 필요.
    """
    try:
        new_model = await VehicleModelService.create_vehicle_model(
            db=db,
            manufacturer_id=request.manufacturer_id,
            model_group=request.model_group,
            model_detail=request.model_detail,
            vehicle_class=request.vehicle_class,
            start_year=request.start_year,
            end_year=request.end_year,
            is_active=request.is_active
        )
        # 제조사 정보 포함하여 응답
        manufacturer = await ManufacturerService.get_manufacturer(db, request.manufacturer_id)
        response_data = {
            "id": new_model.id,
            "manufacturer_id": new_model.manufacturer_id,
            "manufacturer_name": manufacturer.name if manufacturer else None,
            "manufacturer_origin": manufacturer.origin if manufacturer else None,
            "model_group": new_model.model_group,
            "model_detail": new_model.model_detail,
            "vehicle_class": new_model.vehicle_class,
            "start_year": new_model.start_year,
            "end_year": new_model.end_year,
            "is_active": new_model.is_active,
            "created_at": new_model.created_at,
            "updated_at": new_model.updated_at,
        }
        return StandardResponse(success=True, data=response_data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 생성 중 오류 발생: {str(e)}")


@router.get("/vehicle-models/{model_id}", response_model=StandardResponse)
async def get_vehicle_model_detail(
    model_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    특정 차량 모델 상세 정보를 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        model_uuid = uuid.UUID(model_id)
        model = await VehicleModelService.get_vehicle_model(db, model_uuid)
        if not model:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 모델을 찾을 수 없습니다.")
        
        # 제조사 정보 포함
        manufacturer = await ManufacturerService.get_manufacturer(db, model.manufacturer_id)
        response_data = {
            "id": model.id,
            "manufacturer_id": model.manufacturer_id,
            "manufacturer_name": manufacturer.name if manufacturer else None,
            "manufacturer_origin": manufacturer.origin if manufacturer else None,
            "model_group": model.model_group,
            "model_detail": model.model_detail,
            "vehicle_class": model.vehicle_class,
            "start_year": model.start_year,
            "end_year": model.end_year,
            "is_active": model.is_active,
            "created_at": model.created_at,
            "updated_at": model.updated_at,
        }
        return StandardResponse(success=True, data=response_data)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="유효하지 않은 차량 모델 ID 형식입니다.")
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 조회 중 오류 발생: {str(e)}")


@router.patch("/vehicle-models/{model_id}", response_model=StandardResponse)
async def update_vehicle_model(
    model_id: str,
    request: VehicleModelUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 모델 정보를 업데이트합니다.
    관리자 권한 필요.
    """
    try:
        model_uuid = uuid.UUID(model_id)
        updated_model = await VehicleModelService.update_vehicle_model(
            db=db,
            model_id=model_uuid,
            manufacturer_id=request.manufacturer_id,
            model_group=request.model_group,
            model_detail=request.model_detail,
            vehicle_class=request.vehicle_class,
            start_year=request.start_year,
            end_year=request.end_year,
            is_active=request.is_active
        )
        if not updated_model:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 모델을 찾을 수 없습니다.")
        
        # 제조사 정보 포함
        manufacturer = await ManufacturerService.get_manufacturer(db, updated_model.manufacturer_id)
        response_data = {
            "id": updated_model.id,
            "manufacturer_id": updated_model.manufacturer_id,
            "manufacturer_name": manufacturer.name if manufacturer else None,
            "manufacturer_origin": manufacturer.origin if manufacturer else None,
            "model_group": updated_model.model_group,
            "model_detail": updated_model.model_detail,
            "vehicle_class": updated_model.vehicle_class,
            "start_year": updated_model.start_year,
            "end_year": updated_model.end_year,
            "is_active": updated_model.is_active,
            "created_at": updated_model.created_at,
            "updated_at": updated_model.updated_at,
        }
        return StandardResponse(success=True, data=response_data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 업데이트 중 오류 발생: {str(e)}")


@router.delete("/vehicle-models/{model_id}", response_model=StandardResponse)
async def delete_vehicle_model(
    model_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 모델을 삭제합니다 (soft delete).
    관리자 권한 필요.
    """
    try:
        model_uuid = uuid.UUID(model_id)
        success = await VehicleModelService.delete_vehicle_model(db, model_uuid)
        if not success:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="차량 모델을 찾을 수 없습니다.")
        return StandardResponse(success=True, data={"message": "차량 모델이 성공적으로 삭제되었습니다."})
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 삭제 중 오류 발생: {str(e)}")


@router.get("/vehicle-models", response_model=StandardResponse)
async def list_vehicle_models(
    manufacturer_id: Optional[str] = Query(None, description="제조사 ID 필터"),
    origin: Optional[str] = Query(None, description="국산/수입 필터 (domestic, imported)"),
    vehicle_class: Optional[str] = Query(None, description="차량 등급 필터"),
    model_group: Optional[str] = Query(None, description="모델 그룹 필터"),
    model_detail: Optional[str] = Query(None, description="모델 상세 필터"),
    search: Optional[str] = Query(None, description="검색어 (제조사명, 모델명)"),
    is_active: Optional[bool] = Query(None, description="활성화 여부 필터"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    차량 모델 목록을 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        manufacturer_uuid = uuid.UUID(manufacturer_id) if manufacturer_id else None
        models_data = await VehicleModelService.list_vehicle_models(
            db=db,
            manufacturer_id=manufacturer_uuid,
            origin=origin,
            vehicle_class=vehicle_class,
            model_group=model_group,
            model_detail=model_detail,
            search=search,
            is_active=is_active,
            page=page,
            limit=limit
        )
        return StandardResponse(success=True, data=models_data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 목록 조회 중 오류 발생: {str(e)}")


@router.post("/vehicle-models/sync", response_model=StandardResponse)
async def sync_vehicle_models(
    request: VehicleModelSyncRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    차량 모델 데이터를 일괄 동기화합니다.
    관리자 권한 필요.
    """
    try:
        sync_data = [item.model_dump() for item in request.items]
        result = await VehicleModelService.sync_vehicle_models(db, sync_data)
        return StandardResponse(success=True, data=result)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"차량 모델 동기화 중 오류 발생: {str(e)}")


@router.get("/prices", response_model=StandardResponse)
async def list_price_policies(
    origin: Optional[str] = Query(None, description="국산/수입 필터", pattern="^(domestic|imported)$"),
    vehicle_class: Optional[str] = Query(None, description="차량 등급 필터", pattern="^(compact|small|mid|large|suv|sports|supercar)$"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(100, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    가격 정책 목록 조회 API
    
    국산/수입, 차량 등급별로 필터링하여 가격 정책 목록을 조회합니다.
    """
    try:
        result = await PricePolicyService.list_price_policies(
            db=db,
            origin=origin,
            vehicle_class=vehicle_class,
            page=page,
            limit=limit
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"가격 정책 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/prices/{policy_id}", response_model=StandardResponse)
async def get_price_policy(
    policy_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    가격 정책 상세 조회 API
    
    특정 가격 정책의 상세 정보를 조회합니다.
    """
    try:
        result = await PricePolicyService.get_price_policy(
            db=db,
            policy_id=policy_id
        )
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="가격 정책을 찾을 수 없습니다"
            )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"가격 정책 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/prices", response_model=StandardResponse)
async def create_price_policy(
    request: PricePolicyCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    가격 정책 생성 API
    
    차량 등급별 추가 요금 정책을 생성합니다.
    Redis 캐시가 자동으로 무효화됩니다.
    """
    try:
        result = await PricePolicyService.create_price_policy(
            db=db,
            origin=request.origin,
            vehicle_class=request.vehicle_class,
            add_amount=request.add_amount
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"가격 정책 생성 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/prices/{policy_id}", response_model=StandardResponse)
async def update_price_policy(
    policy_id: str,
    request: PricePolicyUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    가격 정책 수정 API
    
    가격 정책의 추가 금액을 수정합니다.
    Redis 캐시가 자동으로 무효화됩니다.
    """
    try:
        result = await PricePolicyService.update_price_policy(
            db=db,
            policy_id=policy_id,
            add_amount=request.add_amount
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"가격 정책 수정 중 오류가 발생했습니다: {str(e)}"
        )


@router.delete("/prices/{policy_id}", response_model=StandardResponse)
async def delete_price_policy(
    policy_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    가격 정책 삭제 API
    
    가격 정책을 삭제합니다.
    Redis 캐시가 자동으로 무효화됩니다.
    """
    try:
        result = await PricePolicyService.delete_price_policy(
            db=db,
            policy_id=policy_id
        )
        
        return StandardResponse(
            success=True,
            data={"deleted": result},
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"가격 정책 삭제 중 오류가 발생했습니다: {str(e)}"
        )


# ==================== 서비스 지역 관리 API ====================

@router.get("/regions", response_model=StandardResponse)
async def list_service_regions(
    province: Optional[str] = Query(None, description="상위 지역 필터"),
    city: Optional[str] = Query(None, description="하위 지역 필터"),
    is_active: Optional[bool] = Query(None, description="활성 상태 필터"),
    search: Optional[str] = Query(None, description="검색어 (상위/하위 지역)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(100, ge=1, le=100, description="페이지 크기"),
    hierarchy: bool = Query(False, description="계층 구조로 반환"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    서비스 지역 목록 조회 API
    
    계층 구조 또는 평면 목록으로 서비스 지역을 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        if hierarchy:
            result = await ServiceRegionService.list_service_regions_hierarchy(
                db=db,
                is_active=is_active
            )
        else:
            result = await ServiceRegionService.list_service_regions(
                db=db,
                province=province,
                city=city,
                is_active=is_active,
                search=search,
                page=page,
                limit=limit
            )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"서비스 지역 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/regions/{region_id}", response_model=StandardResponse)
async def get_service_region(
    region_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    서비스 지역 상세 조회 API
    
    특정 서비스 지역의 상세 정보를 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await ServiceRegionService.get_service_region(
            db=db,
            region_id=region_id
        )
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="서비스 지역을 찾을 수 없습니다"
            )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"서비스 지역 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/regions", response_model=StandardResponse)
async def create_service_region(
    request: ServiceRegionCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    서비스 지역 생성 API
    
    새 서비스 지역을 생성합니다.
    Redis 캐시가 자동으로 무효화됩니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await ServiceRegionService.create_service_region(
            db=db,
            province=request.province,
            province_code=request.province_code,
            city=request.city,
            city_code=request.city_code,
            extra_fee=request.extra_fee,
            is_active=request.is_active
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"서비스 지역 생성 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/regions/{region_id}", response_model=StandardResponse)
async def update_service_region(
    region_id: str,
    request: ServiceRegionUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    서비스 지역 수정 API
    
    서비스 지역 정보를 수정합니다.
    Redis 캐시가 자동으로 무효화됩니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await ServiceRegionService.update_service_region(
            db=db,
            region_id=region_id,
            province=request.province,
            province_code=request.province_code,
            city=request.city,
            city_code=request.city_code,
            extra_fee=request.extra_fee,
            is_active=request.is_active
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"서비스 지역 수정 중 오류가 발생했습니다: {str(e)}"
        )


@router.delete("/regions/{region_id}", response_model=StandardResponse)
async def delete_service_region(
    region_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    서비스 지역 삭제 API
    
    서비스 지역을 삭제합니다.
    활성 신청 건이 있으면 삭제할 수 없습니다.
    Redis 캐시가 자동으로 무효화됩니다.
    관리자 권한 필요.
    """
    try:
        result = await ServiceRegionService.delete_service_region(
            db=db,
            region_id=region_id
        )
        
        return StandardResponse(
            success=True,
            data={"deleted": result},
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"서비스 지역 삭제 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/regions/bulk-update-province", response_model=StandardResponse)
async def bulk_update_province_regions(
    province_code: str = Query(..., description="광역시도 코드 (11, 21, 22 등)"),
    is_active: bool = Query(..., description="활성화 여부"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    광역시도별 서비스 지역 일괄 활성/비활성화 API
    해당 광역시도의 모든 시군구를 일괄로 활성화 또는 비활성화합니다.
    시군구가 없으면 자동으로 생성합니다.
    관리자 권한 필요.
    """
    try:
        result = await ServiceRegionService.bulk_update_province_regions(
            db=db,
            province_code=province_code,
            is_active=is_active
        )
        return StandardResponse(success=True, data=result)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"일괄 업데이트 중 오류 발생: {str(e)}")


@router.get("/regions/province-status/{province_code}", response_model=StandardResponse)
async def get_province_status(
    province_code: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_staff)
):
    """
    광역시도별 활성 지역 수 조회 API
    관리자/직원 권한 필요.
    """
    try:
        result = await ServiceRegionService.get_province_status(db=db, province_code=province_code)
        return StandardResponse(success=True, data=result)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"상태 조회 중 오류 발생: {str(e)}")


@router.get("/inspections", response_model=StandardResponse)
async def get_inspections(
    status: Optional[str] = Query(None, description="상태 필터"),
    region: Optional[str] = Query(None, description="지역 필터"),
    date: Optional[str] = Query(None, description="날짜 필터 (YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    sort_by: str = Query("created_at", description="정렬 기준"),
    sort_order: str = Query("desc", description="정렬 순서"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    신청 목록 조회 API
    
    필터링, 정렬, 페이지네이션을 지원합니다.
    관리자 권한 필요.
    """
    try:
        from datetime import datetime as dt
        
        target_date = None
        if date:
            try:
                target_date = dt.strptime(date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)"
                )
        
        result = await AdminService.get_inspections(
            db=db,
            status=status,
            region=region,
            target_date=target_date,
            page=page,
            limit=limit,
            sort_by=sort_by,
            sort_order=sort_order
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"신청 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/inspections/{inspection_id}", response_model=StandardResponse)
async def get_inspection_detail(
    inspection_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    신청 상세 조회 API (관리자용)
    
    관리자가 신청 상세 정보를 조회합니다.
    관리자 권한 필요.
    """
    try:
        from app.services.inspection_service import InspectionService
        
        # Inspection 조회
        from app.models.inspection import Inspection
        from sqlalchemy import select
        
        inspection_result = await db.execute(
            select(Inspection).where(Inspection.id == inspection_id)
        )
        inspection = inspection_result.scalar_one_or_none()
        
        if not inspection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="진단 신청을 찾을 수 없습니다"
            )
        
        # 관리자용 상세 정보 조회 (user_id는 inspection의 user_id 사용)
        result = await InspectionService.get_inspection_detail(
            db=db,
            inspection_id=inspection_id,
            user_id=str(inspection.user_id)
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"신청 상세 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/inspections/{inspection_id}/assign", response_model=StandardResponse)
async def assign_inspector(
    inspection_id: str,
    request: InspectionAssignRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    강제 배정 API
    
    관리자가 직접 기사를 배정합니다.
    Inspection 상태를 'assigned'로 변경합니다.
    """
    try:
        result = await AdminService.assign_inspector(
            db=db,
            inspection_id=inspection_id,
            inspector_id=request.inspector_id
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"기사 배정 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/inspections/{inspection_id}/status", response_model=StandardResponse)
async def update_inspection_status(
    inspection_id: str,
    status: str = Query(..., description="새 상태"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    신청 상태 변경 API
    
    관리자가 신청 상태를 변경합니다.
    관리자 권한 필요.
    """
    try:
        from app.models.inspection import Inspection
        from sqlalchemy import select
        
        # Inspection 조회
        result = await db.execute(
            select(Inspection).where(Inspection.id == inspection_id)
        )
        inspection = result.scalar_one_or_none()
        
        if not inspection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="진단 신청을 찾을 수 없습니다"
            )
        
        # 유효한 상태인지 확인
        valid_statuses = ["requested", "paid", "assigned", "in_progress", "completed", "sent", "cancelled"]
        if status not in valid_statuses:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"유효하지 않은 상태입니다: {status}"
            )
        
        # 상태 변경
        inspection.status = status
        await db.commit()
        await db.refresh(inspection)
        
        return StandardResponse(
            success=True,
            data={
                "inspection_id": str(inspection.id),
                "status": inspection.status
            },
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"상태 변경 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/reports/{inspection_id}/approve", response_model=StandardResponse)
async def approve_report(
    inspection_id: str,
    feedback: Optional[str] = Query(None, description="피드백"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    레포트 승인 API
    
    관리자가 제출된 레포트를 승인합니다.
    관리자 권한 필요.
    """
    try:
        from app.models.inspection import Inspection
        from app.models.inspection_report import InspectionReport
        from sqlalchemy import select
        
        # Inspection 조회
        result = await db.execute(
            select(Inspection).where(Inspection.id == inspection_id)
        )
        inspection = result.scalar_one_or_none()
        
        if not inspection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="진단 신청을 찾을 수 없습니다"
            )
        
        # InspectionReport 조회
        report_result = await db.execute(
            select(InspectionReport).where(InspectionReport.inspection_id == inspection_id)
        )
        report = report_result.scalar_one_or_none()
        
        if not report:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="레포트를 찾을 수 없습니다"
            )
        
        # 레포트 상태를 승인으로 변경
        report.status = "approved"
        inspection.status = "sent"  # Inspection 상태도 발송완료로 변경
        
        await db.commit()
        await db.refresh(report)
        await db.refresh(inspection)
        
        # 알림 트리거 (고객에게 레포트 발송 알림)
        from app.services.notification_trigger_service import NotificationTriggerService
        await NotificationTriggerService.trigger_report_approved(
            db=db,
            inspection_id=inspection_id,
            user_id=str(inspection.user_id)
        )
        
        return StandardResponse(
            success=True,
            data={
                "inspection_id": str(inspection.id),
                "report_id": str(report.id),
                "status": "approved"
            },
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"레포트 승인 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/reports/{inspection_id}/reject", response_model=StandardResponse)
async def reject_report(
    inspection_id: str,
    feedback: Optional[str] = Query(None, description="반려 사유"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    레포트 반려 API
    
    관리자가 제출된 레포트를 반려합니다.
    관리자 권한 필요.
    """
    try:
        from app.models.inspection import Inspection
        from app.models.inspection_report import InspectionReport
        from sqlalchemy import select
        
        # Inspection 조회
        result = await db.execute(
            select(Inspection).where(Inspection.id == inspection_id)
        )
        inspection = result.scalar_one_or_none()
        
        if not inspection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="진단 신청을 찾을 수 없습니다"
            )
        
        # InspectionReport 조회
        report_result = await db.execute(
            select(InspectionReport).where(InspectionReport.inspection_id == inspection_id)
        )
        report = report_result.scalar_one_or_none()
        
        if not report:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="레포트를 찾을 수 없습니다"
            )
        
        # 레포트 상태를 반려로 변경
        report.status = "rejected"
        if feedback:
            # 피드백을 inspector_comment에 추가하거나 별도 필드에 저장
            # 현재는 간단히 상태만 변경
            pass
        
        await db.commit()
        await db.refresh(report)
        
        # 알림 트리거 (기사에게 수정 요청 알림)
        if inspection.inspector_id:
            from app.services.notification_trigger_service import NotificationTriggerService
            await NotificationTriggerService.trigger_report_rejected(
                db=db,
                inspection_id=inspection_id,
                inspector_id=str(inspection.inspector_id),
                feedback=feedback or ""
            )
        
        return StandardResponse(
            success=True,
            data={
                "inspection_id": str(inspection.id),
                "report_id": str(report.id),
                "status": "rejected"
            },
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"레포트 반려 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/settlements/calculate", response_model=StandardResponse)
async def calculate_settlements(
    request: SettlementCalculateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 집계 API
    
    특정 날짜 완료 건에 대한 정산을 계산하고 Settlement 레코드를 생성합니다.
    관리자 권한 필요.
    """
    try:
        result = await AdminService.calculate_settlements(
            db=db,
            target_date=request.target_date
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 집계 중 오류가 발생했습니다: {str(e)}"
        )


# ==================== 유저 관리 API ====================

@router.post("/users", response_model=StandardResponse)
async def create_user(
    request: UserCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"], require_admin_for_admin_role=True))
):
    """
    유저 생성 API
    
    관리자가 새 유저를 생성합니다.
    관리자/직원 권한 필요.
    """
    try:
        # commission_rate를 0~1 범위로 변환 (프론트엔드는 0~100으로 전송)
        commission_rate = None
        if request.commission_rate is not None:
            # 0~100 범위를 0~1 범위로 변환
            commission_rate = float(request.commission_rate) / 100.0
            if commission_rate < 0 or commission_rate > 1:
                raise ValueError("수수료율은 0~100 사이여야 합니다")
        
        result = await UserService.create_user(
            db=db,
            role=request.role,
            name=request.name,
            phone=request.phone,
            email=request.email,
            password=request.password,
            region_ids=request.region_ids,
            level=request.level,
            commission_rate=commission_rate,
            status=request.status
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"유저 생성 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/users/{user_id}", response_model=StandardResponse)
async def get_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    유저 상세 조회 API
    
    관리자가 유저 상세 정보를 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await UserService.get_user(db=db, user_id=user_id)
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="유저를 찾을 수 없습니다"
            )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"유저 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/users/{user_id}", response_model=StandardResponse)
async def update_user(
    user_id: str,
    request: UserUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    유저 정보 수정 API
    
    관리자가 유저 정보를 수정합니다.
    관리자/직원 권한 필요.
    """
    try:
        # commission_rate를 0~1 범위로 변환 (프론트엔드는 0~100으로 전송)
        commission_rate = None
        if request.commission_rate is not None:
            # 0~100 범위를 0~1 범위로 변환
            commission_rate = float(request.commission_rate) / 100.0
            if commission_rate < 0 or commission_rate > 1:
                raise ValueError("수수료율은 0~100 사이여야 합니다")
        
        result = await UserService.update_user(
            db=db,
            user_id=user_id,
            name=request.name,
            email=request.email,
            phone=request.phone,
            password=request.password,
            region_ids=request.region_ids,
            level=request.level,
            commission_rate=commission_rate,
            status=request.status
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"유저 수정 중 오류가 발생했습니다: {str(e)}"
        )


@router.delete("/users/{user_id}", response_model=StandardResponse)
async def delete_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    유저 삭제 API (Soft Delete)
    
    관리자가 유저를 삭제합니다. 실제로는 상태를 inactive로 변경합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await UserService.delete_user(db=db, user_id=user_id)
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"유저 삭제 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/users", response_model=StandardResponse)
async def list_users(
    role: Optional[str] = Query(None, description="역할 필터 (admin, user, partner)"),
    user_status: Optional[str] = Query(None, alias="status", description="계정 상태 필터 (active, inactive, suspended)"),
    level: Optional[int] = Query(None, description="등급 필터 (기사용)"),
    search: Optional[str] = Query(None, description="검색어 (이름, 이메일)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    유저 목록 조회 API
    
    필터링, 검색, 페이지네이션을 지원합니다.
    관리자/직원 권한 필요.
    """
    try:
        offset = (page - 1) * limit
        
        result = await UserService.list_users(
            db=db,
            role=role,
            status=user_status,
            level=level,
            search=search,
            offset=offset,
            limit=limit
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        import traceback
        print(f"Error in list_users: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"유저 목록 조회 중 오류가 발생했습니다: {repr(e)}"
        )


# ==================== 유저 등급/역할/상태 관리 API ====================

@router.patch("/users/{user_id}/level", response_model=StandardResponse)
async def update_user_level(
    user_id: str,
    request: UserLevelUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    기사 등급 변경 API
    
    기사의 등급을 변경합니다 (1~5).
    관리자/직원 권한 필요.
    """
    try:
        result = await UserService.update_user_level(
            db=db,
            user_id=user_id,
            level=request.level
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"등급 변경 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/users/{user_id}/commission", response_model=StandardResponse)
async def update_user_commission(
    user_id: str,
    request: UserCommissionUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    수수료율 변경 API
    
    기사의 수수료율을 변경합니다 (0~100%).
    관리자/직원 권한 필요.
    """
    try:
        result = await UserService.update_user_commission(
            db=db,
            user_id=user_id,
            commission_rate=float(request.commission_rate)
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"수수료율 변경 중 오류가 발생했습니다: {str(e)}"
        )




@router.patch("/users/{user_id}/role", response_model=StandardResponse)
async def update_user_role(
    user_id: str,
    request: UserRoleUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only())
):
    """
    역할 변경 API
    
    유저의 역할을 변경합니다.
    - admin 역할 부여는 admin만 가능
    - 자기 자신의 역할 변경 불가
    관리자 권한 필요.
    """
    try:
        result = await UserService.update_user_role(
            db=db,
            user_id=user_id,
            new_role=request.role,
            current_user_id=str(current_user.id)
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"역할 변경 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/users/{user_id}/status", response_model=StandardResponse)
async def update_user_status(
    user_id: str,
    request: UserStatusUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    계정 상태 변경 API
    
    유저의 계정 상태를 변경합니다 (active/inactive/suspended).
    관리자/직원 권한 필요.
    """
    try:
        result = await UserService.update_user_status(
            db=db,
            user_id=user_id,
            new_status=request.status
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"상태 변경 중 오류가 발생했습니다: {str(e)}"
        )


# ==================== 패키지 관리 API ====================

@router.post("/packages", response_model=StandardResponse)
async def create_package(
    request: PackageCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    패키지 생성 API
    
    관리자가 새 패키지를 생성합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await PackageService.create_package(
            db=db,
            name=request.name,
            base_price=request.base_price,
            included_items=request.included_items
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"패키지 생성 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/packages/{package_id}", response_model=StandardResponse)
async def get_package(
    package_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    패키지 상세 조회 API
    
    관리자가 패키지 상세 정보를 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await PackageService.get_package(db=db, package_id=package_id)
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="패키지를 찾을 수 없습니다"
            )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"패키지 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/packages/{package_id}", response_model=StandardResponse)
async def update_package(
    package_id: str,
    request: PackageUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    패키지 수정 API
    
    관리자가 패키지 정보를 수정합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await PackageService.update_package(
            db=db,
            package_id=package_id,
            name=request.name,
            base_price=request.base_price,
            included_items=request.included_items,
            is_active=request.is_active
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"패키지 수정 중 오류가 발생했습니다: {str(e)}"
        )


@router.delete("/packages/{package_id}", response_model=StandardResponse)
async def delete_package(
    package_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    패키지 삭제 API (Soft Delete)
    
    관리자가 패키지를 삭제합니다. 실제로는 is_active를 False로 변경합니다.
    활성 신청 건이 있으면 삭제할 수 없습니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await PackageService.delete_package(db=db, package_id=package_id)
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"패키지 삭제 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/packages", response_model=StandardResponse)
async def list_packages(
    search: Optional[str] = Query(None, description="검색어 (패키지 이름)"),
    is_active: Optional[bool] = Query(None, description="활성화 여부 필터"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    패키지 목록 조회 API
    
    필터링, 검색, 페이지네이션을 지원합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await PackageService.list_packages(
            db=db,
            search=search,
            is_active=is_active,
            page=page,
            limit=limit
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"패키지 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


# ============================================
# 정산 관리 API
# ============================================

@router.get("/settlements", response_model=StandardResponse)
async def list_settlements(
    inspector_id: Optional[str] = Query(None, description="기사 ID (필터링)"),
    status: Optional[str] = Query(None, description="정산 상태 (pending, completed)"),
    start_date: Optional[date] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="종료일 (YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    page_size: int = Query(20, ge=1, le=100, description="페이지 크기"),
    sort_by: str = Query("settle_date", description="정렬 기준 (settle_date, settle_amount, created_at)"),
    sort_order: str = Query("desc", description="정렬 순서 (asc, desc)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 내역 목록 조회 API
    
    필터링, 정렬, 페이지네이션을 지원합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.get_settlements(
            db=db,
            inspector_id=inspector_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 내역 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/settlements/{settlement_id}", response_model=StandardResponse)
async def get_settlement_detail(
    settlement_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 상세 내역 조회 API
    
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.get_settlement_detail(
            db=db,
            settlement_id=settlement_id
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 상세 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/settlements/inspector/{inspector_id}", response_model=StandardResponse)
async def get_inspector_settlements(
    inspector_id: str,
    status: Optional[str] = Query(None, description="정산 상태 (pending, completed)"),
    start_date: Optional[date] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="종료일 (YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    page_size: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    기사별 정산 내역 조회 API
    
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.get_settlements(
            db=db,
            inspector_id=inspector_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
            page=page,
            page_size=page_size
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"기사별 정산 내역 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/settlements/summary", response_model=StandardResponse)
async def get_settlement_summary(
    start_date: Optional[date] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="종료일 (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 요약 정보 조회 API
    
    일/주/월 단위 정산 예정액 및 기사별 정산 현황을 조회합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.get_settlement_summary(
            db=db,
            start_date=start_date,
            end_date=end_date
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 요약 조회 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/settlements/calculate", response_model=StandardResponse)
async def calculate_settlements(
    request: SettlementCalculateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 집계 실행 API (수동)
    
    지정된 날짜에 완료된 진단 건에 대한 정산을 집계합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await AdminService.calculate_settlements(
            db=db,
            target_date=request.target_date
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 집계 중 오류가 발생했습니다: {str(e)}"
        )


@router.patch("/settlements/{settlement_id}/status", response_model=StandardResponse)
async def update_settlement_status(
    settlement_id: str,
    request: SettlementStatusUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 상태 변경 API
    
    정산 상태를 변경합니다 (pending → completed).
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.update_settlement_status(
            db=db,
            settlement_id=settlement_id,
            status=request.status
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 상태 변경 중 오류가 발생했습니다: {str(e)}"
        )


@router.post("/settlements/bulk-update", response_model=StandardResponse)
async def bulk_update_settlement_status(
    request: SettlementBulkUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 일괄 상태 변경 API
    
    여러 정산 건의 상태를 일괄 변경합니다.
    관리자/직원 권한 필요.
    """
    try:
        result = await SettlementService.bulk_update_settlement_status(
            db=db,
            settlement_ids=request.settlement_ids,
            status=request.status
        )
        
        return StandardResponse(
            success=True,
            data=result,
            error=None
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 일괄 상태 변경 중 오류가 발생했습니다: {str(e)}"
        )


@router.get("/settlements/export")
async def export_settlements(
    inspector_id: Optional[str] = Query(None, description="기사 ID (필터링)"),
    status: Optional[str] = Query(None, description="정산 상태 (pending, completed)"),
    start_date: Optional[date] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="종료일 (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    정산 내역 엑셀 다운로드 API
    
    정산 내역을 엑셀 파일로 다운로드합니다 (세무처리용).
    관리자/직원 권한 필요.
    """
    try:
        # openpyxl import (조건부)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="엑셀 다운로드 기능을 사용할 수 없습니다. openpyxl 라이브러리 설치가 필요합니다."
            )
        
        # 정산 내역 조회 (대량 데이터)
        result = await SettlementService.get_settlements(
            db=db,
            inspector_id=inspector_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
            page=1,
            page_size=10000  # 대량 데이터 조회
        )
        
        settlements = result["settlements"]
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "정산 내역"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = [
            "No",
            "정산 ID",
            "기사명",
            "진단 ID",
            "고객 결제금액",
            "수수료율",
            "정산액",
            "정산 상태",
            "정산일",
            "생성일",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_num, settlement in enumerate(settlements, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)  # No
            ws.cell(row=row_num, column=2, value=settlement["id"])  # 정산 ID
            ws.cell(row=row_num, column=3, value=settlement.get("inspector_name") or "-")  # 기사명
            ws.cell(row=row_num, column=4, value=settlement["inspection_id"])  # 진단 ID
            ws.cell(row=row_num, column=5, value=settlement["total_sales"])  # 고객 결제금액
            ws.cell(row=row_num, column=6, value=f"{settlement['fee_rate'] * 100:.1f}%")  # 수수료율
            ws.cell(row=row_num, column=7, value=settlement["settle_amount"])  # 정산액
            ws.cell(row=row_num, column=8, value="정산완료" if settlement["status"] == "completed" else "미정산")  # 정산 상태
            ws.cell(row=row_num, column=9, value=settlement["settle_date"])  # 정산일
            ws.cell(row=row_num, column=10, value=settlement["created_at"])  # 생성일
        
        # 컬럼 너비 자동 조정
        column_widths = [6, 36, 15, 36, 15, 12, 15, 12, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # 숫자 형식 적용 (금액 컬럼)
        for row_num in range(2, len(settlements) + 2):
            # 고객 결제금액 (컬럼 E)
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            # 정산액 (컬럼 G)
            ws.cell(row=row_num, column=7).number_format = '#,##0'
        
        # 메모리 버퍼에 엑셀 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명 생성
        from datetime import datetime
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f"정산내역_{date_str}.xlsx"
        if start_date and end_date:
            filename = f"정산내역_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        # StreamingResponse로 반환
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"정산 내역 다운로드 중 오류가 발생했습니다: {str(e)}"
        )


# ============================================
# 리뷰 관리 API
# ============================================

@router.get("/reviews", response_model=StandardResponse)
async def list_reviews(
    rating: Optional[int] = Query(None, description="별점 필터 (1-5)"),
    is_hidden: Optional[bool] = Query(None, description="숨김 여부 필터"),
    page: int = Query(1, ge=1, description="페이지 번호"),
    limit: int = Query(20, ge=1, le=100, description="페이지 크기"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    리뷰 목록 조회 API
    """
    try:
        offset = (page - 1) * limit
        result = await ReviewService.get_reviews(
            db=db,
            skip=offset,
            limit=limit,
            rating=rating,
            is_hidden=is_hidden
        )
        
        return StandardResponse(
            success=True,
            data={
                "items": [ReviewResponse.model_validate(item) for item in result["items"]],
                "total": result["total"],
                "page": page,
                "limit": limit,
                "total_pages": (result["total"] + limit - 1) // limit
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"리뷰 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )

@router.patch("/reviews/{review_id}/visibility", response_model=StandardResponse)
async def update_review_visibility(
    review_id: str,
    request: ReviewUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    리뷰 숨김 상태 변경 API
    """
    try:
        review_uuid = uuid.UUID(review_id)
        if request.is_hidden is None:
             raise ValueError("is_hidden 필드가 필요합니다.")
             
        review = await ReviewService.update_visibility(
            db=db,
            review_id=review_uuid,
            is_hidden=request.is_hidden
        )
        
        if not review:
            raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다.")
            
        return StandardResponse(success=True, data=ReviewResponse.model_validate(review))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"리뷰 상태 변경 중 오류: {str(e)}")


# ============================================
# FAQ 관리 API
# ============================================

@router.get("/faqs", response_model=StandardResponse)
async def list_faqs(
    category: Optional[str] = Query(None, description="카테고리 필터"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "staff"]))
):
    """
    FAQ 목록 조회 API
    """
    try:
        faqs = await FAQService.get_faqs(db=db, category=category)
        return StandardResponse(
            success=True,
            data={
                "items": [FAQResponse.model_validate(faq) for faq in faqs],
                "total": len(faqs)
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"FAQ 목록 조회 중 오류: {str(e)}")

@router.post("/faqs", response_model=StandardResponse)
async def create_faq(
    request: FAQCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    FAQ 생성 API
    """
    try:
        faq = await FAQService.create_faq(
            db=db,
            category=request.category,
            question=request.question,
            answer=request.answer,
            is_active=request.is_active,
            display_order=request.display_order
        )
        return StandardResponse(success=True, data=FAQResponse.model_validate(faq))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"FAQ 생성 중 오류: {str(e)}")

@router.patch("/faqs/{faq_id}", response_model=StandardResponse)
async def update_faq(
    faq_id: str,
    request: FAQUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    FAQ 수정 API
    """
    try:
        faq_uuid = uuid.UUID(faq_id)
        update_data = request.model_dump(exclude_unset=True)
        if not update_data:
            raise ValueError("변경할 데이터가 없습니다.")
            
        faq = await FAQService.update_faq(db=db, faq_id=faq_uuid, **update_data)
        if not faq:
            raise HTTPException(status_code=404, detail="FAQ를 찾을 수 없습니다.")
            
        return StandardResponse(success=True, data=FAQResponse.model_validate(faq))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"FAQ 수정 중 오류: {str(e)}")

@router.delete("/faqs/{faq_id}", response_model=StandardResponse)
async def delete_faq(
    faq_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_only)
):
    """
    FAQ 삭제 API
    """
    try:
        faq_uuid = uuid.UUID(faq_id)
        success = await FAQService.delete_faq(db=db, faq_id=faq_uuid)
        if not success:
            raise HTTPException(status_code=404, detail="FAQ를 찾을 수 없습니다.")
            
        return StandardResponse(success=True, data={"message": "FAQ가 삭제되었습니다."})
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"FAQ 삭제 중 오류: {str(e)}")
